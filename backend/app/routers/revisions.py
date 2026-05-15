"""Полноценный модуль ревизии (инвентаризации).

Жизненный цикл:
1. Owner создаёт активную ревизию (POST /revisions). Активная одна на магазин.
2. Warehouse и Owner вносят/обновляют фактическое количество товара
   (POST /revisions/{id}/items, по product_id, upsert).
3. Owner завершает (POST /revisions/{id}/complete) — на каждой строке с
   разницей создаётся StockMovement (in или out), остатки пересчитываются.
4. Завершённую ревизию редактировать нельзя.

Существующая «быстрая ревизия» в /stock/revision остаётся как ad-hoc операция
(одиночный кассир, без сохранения состояния).
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from decimal import Decimal
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Response
from pydantic import BaseModel
from sqlalchemy import and_, case, func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.dependencies import get_current_user, require_role
from app.models.product import Product
from app.models.revision import Revision, RevisionItem
from app.models.stock import StockMovement, StockMovementType
from app.models.user import User

router = APIRouter(prefix="/revisions", tags=["revisions"])


def _zero() -> Decimal:
    return Decimal("0")


# ===================== Pydantic schemas =====================

class RevisionItemIn(BaseModel):
    product_id: int
    actual_qty: Decimal


class RevisionCreate(BaseModel):
    note: str | None = None


# ===================== Helpers =====================

async def _get_revision_or_404(db: AsyncSession, rev_id: int, org_id: int) -> Revision:
    rev = (
        await db.execute(
            select(Revision).where(Revision.id == rev_id, Revision.org_id == org_id)
        )
    ).scalar_one_or_none()
    if not rev:
        raise HTTPException(status_code=404, detail="Ревизия не найдена")
    return rev


async def _stock_balance(db: AsyncSession, product_id: int, org_id: int) -> Decimal:
    """Текущий остаток товара (in − (out + writeoff)). Учитывает is_deleted и весовые."""
    eff_qty = func.coalesce(StockMovement.quantity_decimal, StockMovement.quantity)
    in_q = (
        await db.execute(
            select(func.coalesce(func.sum(eff_qty), 0)).where(
                StockMovement.product_id == product_id,
                StockMovement.org_id == org_id,
                StockMovement.type == StockMovementType.in_stock,
                StockMovement.is_deleted.is_(False),
            )
        )
    ).scalar_one()
    out_q = (
        await db.execute(
            select(func.coalesce(func.sum(eff_qty), 0)).where(
                StockMovement.product_id == product_id,
                StockMovement.org_id == org_id,
                StockMovement.type.in_([StockMovementType.out, StockMovementType.writeoff]),
                StockMovement.is_deleted.is_(False),
            )
        )
    ).scalar_one()
    return Decimal(str(in_q or 0)) - Decimal(str(out_q or 0))


def _serialize_revision(r: Revision) -> dict:
    return {
        "id": r.id,
        "status": r.status,
        "created_by": r.created_by,
        "created_at": r.created_at.isoformat() if r.created_at else None,
        "completed_at": r.completed_at.isoformat() if r.completed_at else None,
        "completed_by": r.completed_by,
        "note": r.note,
    }


# ===================== Endpoints =====================

@router.get("")
async def list_revisions(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> list[dict]:
    """Последние 30 ревизий организации (с именами авторов)."""
    rows = list(
        (
            await db.execute(
                select(Revision)
                .where(Revision.org_id == user.org_id)
                .order_by(Revision.id.desc())
                .limit(30)
            )
        ).scalars().all()
    )
    user_ids = {r.created_by for r in rows} | {r.completed_by for r in rows if r.completed_by}
    users = (
        {
            u.id: u.name
            for u in (
                await db.execute(select(User).where(User.id.in_(user_ids)))
            ).scalars().all()
        }
        if user_ids
        else {}
    )
    return [
        {
            **_serialize_revision(r),
            "created_by_name": users.get(r.created_by),
            "completed_by_name": users.get(r.completed_by) if r.completed_by else None,
        }
        for r in rows
    ]


@router.get("/active")
async def get_active(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict:
    """Текущая активная ревизия. {revision: null} если нет."""
    rev = (
        await db.execute(
            select(Revision).where(Revision.org_id == user.org_id, Revision.status == "active")
        )
    ).scalar_one_or_none()
    return {"revision": _serialize_revision(rev) if rev else None}


@router.post("", status_code=201)
async def create_revision(
    payload: RevisionCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_role("owner")),
) -> dict:
    """Только owner. 409 если уже есть активная."""
    rev = Revision(
        org_id=user.org_id,
        created_by=user.id,
        created_at=datetime.now(timezone.utc),
        status="active",
        note=payload.note,
    )
    db.add(rev)
    try:
        await db.commit()
    except IntegrityError:
        await db.rollback()
        raise HTTPException(status_code=409, detail="Уже есть активная ревизия")
    await db.refresh(rev)
    return _serialize_revision(rev)


@router.get("/{revision_id}")
async def get_revision_details(
    revision_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict:
    """Ревизия + все её позиции + расчётные остатки.
    Для активной — expected_qty из позиции это снимок «на момент ввода факта»."""
    rev = await _get_revision_or_404(db, revision_id, user.org_id)
    items = list(
        (
            await db.execute(
                select(RevisionItem, Product, User.name.label("counted_by_name"))
                .join(Product, Product.id == RevisionItem.product_id)
                .outerjoin(User, User.id == RevisionItem.counted_by)
                .where(RevisionItem.revision_id == revision_id)
                .order_by(RevisionItem.updated_at.desc())
            )
        ).all()
    )
    return {
        **_serialize_revision(rev),
        "items": [
            {
                "id": it.RevisionItem.id,
                "product_id": it.RevisionItem.product_id,
                "product_name": it.Product.name,
                "barcode": it.Product.barcode,
                "expected_qty": str(it.RevisionItem.expected_qty),
                "actual_qty": str(it.RevisionItem.actual_qty),
                "counted_by": it.RevisionItem.counted_by,
                "counted_by_name": it.counted_by_name,
                "updated_at": it.RevisionItem.updated_at.isoformat() if it.RevisionItem.updated_at else None,
            }
            for it in items
        ],
    }


@router.post("/{revision_id}/items")
async def upsert_item(
    revision_id: int,
    payload: RevisionItemIn,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_role("owner", "warehouse")),
) -> dict:
    """Внести/перезаписать фактическое количество товара в ревизии.
    Owner и warehouse. Активная ревизия — иначе 400."""
    rev = await _get_revision_or_404(db, revision_id, user.org_id)
    if rev.status != "active":
        raise HTTPException(status_code=400, detail="Ревизия уже завершена, изменять нельзя")

    product = (
        await db.execute(
            select(Product).where(
                Product.id == payload.product_id,
                Product.org_id == user.org_id,
                Product.is_deleted.is_(False),
            )
        )
    ).scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Товар не найден")

    # Штучный товар не может иметь дробное количество — иначе в complete int()
    # потеряет дробную часть, и недостача исчезнет молча.
    is_weighed = getattr(product, "kind", None) == "weighed"
    if not is_weighed and payload.actual_qty != int(payload.actual_qty):
        raise HTTPException(
            status_code=400,
            detail=f"Товар «{product.name}» штучный — количество должно быть целым (без дробной части)",
        )

    existing = (
        await db.execute(
            select(RevisionItem).where(
                RevisionItem.revision_id == revision_id,
                RevisionItem.product_id == payload.product_id,
            )
        )
    ).scalar_one_or_none()
    now = datetime.now(timezone.utc)
    if existing:
        existing.actual_qty = payload.actual_qty
        existing.counted_by = user.id
        existing.updated_at = now
        item = existing
    else:
        # Фиксируем учётный остаток на момент первого ввода — чтобы продажи
        # параллельно с инвентаризацией не размывали ожидаемое количество.
        expected = await _stock_balance(db, payload.product_id, user.org_id)
        item = RevisionItem(
            revision_id=revision_id,
            product_id=payload.product_id,
            expected_qty=expected,
            actual_qty=payload.actual_qty,
            counted_by=user.id,
            created_at=now,
            updated_at=now,
        )
        db.add(item)
    await db.commit()
    await db.refresh(item)
    return {
        "id": item.id,
        "product_id": item.product_id,
        "expected_qty": str(item.expected_qty),
        "actual_qty": str(item.actual_qty),
        "counted_by": item.counted_by,
        "updated_at": item.updated_at.isoformat() if item.updated_at else None,
    }


@router.post("/{revision_id}/complete")
async def complete_revision(
    revision_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_role("owner")),
) -> dict:
    """Owner завершает ревизию.
    На каждой строке с разницей actual − expected создаётся StockMovement:
    - delta > 0 → in_stock (излишек),
    - delta < 0 → out (недостача).
    Reason указывает номер ревизии для трассировки."""
    rev = await _get_revision_or_404(db, revision_id, user.org_id)
    if rev.status != "active":
        raise HTTPException(status_code=400, detail="Ревизия уже завершена")

    items = list(
        (
            await db.execute(
                select(RevisionItem).where(RevisionItem.revision_id == revision_id)
            )
        ).scalars().all()
    )

    # Сначала проверим, что все продукты ещё существуют (не soft-deleted).
    # Если хоть один удалён за время инвентаризации — не завершаем, иначе недостача
    # по нему молча потеряется, и владелец не узнает.
    product_ids = [it.product_id for it in items]
    products = {
        p.id: p
        for p in (
            await db.execute(
                select(Product).where(
                    Product.id.in_(product_ids), Product.is_deleted.is_(False)
                )
            )
        ).scalars().all()
    } if product_ids else {}
    missing = [pid for pid in product_ids if pid not in products]
    if missing:
        raise HTTPException(
            status_code=409,
            detail=(
                f"Нельзя завершить ревизию: товары удалены за время инвентаризации "
                f"(id: {missing}). Восстанови их или удали соответствующие позиции из ревизии."
            ),
        )

    adjustments_in = 0
    adjustments_out = 0
    for it in items:
        delta = (it.actual_qty or _zero()) - (it.expected_qty or _zero())
        if delta == 0:
            continue
        product = products.get(it.product_id)
        is_weighed = getattr(product, "kind", None) == "weighed"
        qty_abs = abs(delta)
        movement = StockMovement(
            org_id=user.org_id,
            product_id=it.product_id,
            quantity=0 if is_weighed else int(qty_abs),
            quantity_decimal=qty_abs if is_weighed else None,
            type=StockMovementType.in_stock if delta > 0 else StockMovementType.out,
            reason=f"Ревизия #{rev.id}: было {it.expected_qty}, стало {it.actual_qty}",
            created_by=user.id,
        )
        db.add(movement)
        if delta > 0:
            adjustments_in += 1
        else:
            adjustments_out += 1

    rev.status = "completed"
    rev.completed_at = datetime.now(timezone.utc)
    rev.completed_by = user.id
    await db.commit()
    return {
        "detail": "Ревизия завершена",
        "adjustments_in": adjustments_in,
        "adjustments_out": adjustments_out,
        "items_total": len(items),
    }


@router.get("/{revision_id}/report")
async def report(
    revision_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict:
    """Сводный отчёт: расхождения и суммы (по закупочной цене Product.purchase_price)."""
    rev = await _get_revision_or_404(db, revision_id, user.org_id)
    rows = list(
        (
            await db.execute(
                select(RevisionItem, Product)
                .join(Product, Product.id == RevisionItem.product_id)
                .where(RevisionItem.revision_id == revision_id)
                .order_by(Product.name.asc())
            )
        ).all()
    )

    items: list[dict] = []
    surplus_value = _zero()  # излишки в сом
    shortage_value = _zero()  # недостачи в сом
    for it, product in rows:
        delta = (it.actual_qty or _zero()) - (it.expected_qty or _zero())
        cost = product.purchase_price or _zero()
        diff_value = delta * cost
        if delta > 0:
            surplus_value += diff_value
        elif delta < 0:
            shortage_value += diff_value  # отрицательная сумма
        items.append({
            "product_id": product.id,
            "product_name": product.name,
            "barcode": product.barcode,
            "expected_qty": str(it.expected_qty),
            "actual_qty": str(it.actual_qty),
            "delta": str(delta),
            "purchase_price": str(cost),
            "diff_value": str(diff_value),
        })

    return {
        **_serialize_revision(rev),
        "summary": {
            "items_total": len(rows),
            "items_with_diff": sum(1 for r in items if Decimal(r["delta"]) != 0),
            "surplus_value": str(surplus_value),
            "shortage_value": str(shortage_value),
            "net_value": str(surplus_value + shortage_value),
        },
        "items": items,
    }


@router.get("/{revision_id}/export")
async def export_excel(
    revision_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> Response:
    """Экспорт отчёта в Excel (.xlsx) через openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    rep = await report(revision_id, db, user)
    wb = Workbook()
    ws = wb.active
    ws.title = f"Ревизия #{revision_id}"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E0E7FF")
    headers = ["Товар", "Штрихкод", "Ожидалось", "Факт", "Расхождение", "Цена закупки", "Сумма расхождения"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, it in enumerate(rep["items"], 2):
        ws.cell(row=i, column=1, value=it["product_name"])
        ws.cell(row=i, column=2, value=it["barcode"])
        ws.cell(row=i, column=3, value=float(it["expected_qty"]))
        ws.cell(row=i, column=4, value=float(it["actual_qty"]))
        ws.cell(row=i, column=5, value=float(it["delta"]))
        ws.cell(row=i, column=6, value=float(it["purchase_price"]))
        ws.cell(row=i, column=7, value=float(it["diff_value"]))

    # Итоги внизу.
    last = len(rep["items"]) + 3
    ws.cell(row=last, column=1, value="ИТОГО").font = bold
    ws.cell(row=last, column=5, value=f"Излишек: {rep['summary']['surplus_value']}").font = bold
    ws.cell(row=last + 1, column=5, value=f"Недостача: {rep['summary']['shortage_value']}").font = bold
    ws.cell(row=last + 2, column=5, value=f"Чистое расхождение: {rep['summary']['net_value']}").font = bold

    # Ширина столбцов.
    widths = [40, 20, 12, 12, 14, 14, 18]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"revision_{revision_id}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
