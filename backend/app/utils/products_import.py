"""Парсер Umag Excel-экспорта в нормализованные строки для импорта.

Известная структура колонок Umag (A..K):
    A  Название товара                  → name (обязательно)
    B  Категория                        → category ('Незаданные' → NULL)
    C  Подкатегория                     (skip)
    D  Штрихкод                         → barcode (EAN-13, 13 цифр; обязательно)
    E  Доп. штрихкоды                   (skip)
    F  Кол-во                           → quantity (int)
    G  Ед. изм                          (skip)
    H  Цена продажи (за ед.)            → sale_price (Decimal)
    I  Сумма по продажам                (skip — вычисляемое)
    J  Закупочная цена                  → purchase_price (Decimal)
    K  Сумма по закупке                 (skip — вычисляемое)

Первая строка — заголовок, пропускается.
"""
from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import IO

from openpyxl import load_workbook


@dataclass
class ImportRow:
    row_index: int            # 1-based, как в Excel (заголовок = 1, первая запись = 2)
    name: str
    barcode: str
    category: str | None
    sale_price: Decimal
    purchase_price: Decimal
    quantity: int


@dataclass
class ImportRowError:
    row_index: int
    reason: str
    raw: dict


def _to_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_decimal(value: object) -> Decimal:
    if value is None or value == "":
        return Decimal("0.00")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    return Decimal(_to_str(value).replace(",", "."))


def _to_int(value: object) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    return int(float(_to_str(value).replace(",", ".")))


def _normalize_barcode(value: object) -> str:
    """Excel часто хранит длинные числовые штрихкоды как float (научная нотация).
    Преобразуем к строке цифр и дополняем до 13 ведущими нулями (EAN-13)."""
    if value is None or value == "":
        return ""
    if isinstance(value, float):
        as_str = str(int(value))
    elif isinstance(value, int):
        as_str = str(value)
    else:
        as_str = _to_str(value)
    digits = "".join(ch for ch in as_str if ch.isdigit())
    if not digits:
        return ""
    if len(digits) > 13:
        return ""  # некорректный — отбросим, выше пометим как ошибку
    return digits.zfill(13)


def _normalize_category(value: object) -> str | None:
    raw = _to_str(value)
    if not raw or raw.lower() == "незаданные":
        return None
    return raw[:100]


def parse_umag_xlsx(stream: IO[bytes]) -> tuple[list[ImportRow], list[ImportRowError]]:
    """Возвращает (валидные_строки, ошибки). Не пишет в БД."""
    workbook = load_workbook(stream, read_only=True, data_only=True)
    sheet = workbook.active
    rows: list[ImportRow] = []
    errors: list[ImportRowError] = []

    for idx, raw in enumerate(sheet.iter_rows(values_only=True), start=1):
        if idx == 1:
            continue  # шапка
        if raw is None or all(cell is None or cell == "" for cell in raw):
            continue  # пустая строка

        cols = list(raw) + [None] * (11 - len(raw))  # дополним до 11 колонок
        name = _to_str(cols[0])
        category = _normalize_category(cols[1])
        barcode = _normalize_barcode(cols[3])

        snapshot = {"A": cols[0], "B": cols[1], "D": cols[3], "F": cols[5], "H": cols[7], "J": cols[9]}

        if not name:
            errors.append(ImportRowError(idx, "пустое название (колонка A)", snapshot))
            continue
        if not barcode:
            errors.append(ImportRowError(idx, "пустой/некорректный штрихкод (колонка D)", snapshot))
            continue

        try:
            sale_price = _to_decimal(cols[7])
            purchase_price = _to_decimal(cols[9])
            quantity = _to_int(cols[5])
        except (InvalidOperation, ValueError) as exc:
            errors.append(ImportRowError(idx, f"невалидное число: {exc}", snapshot))
            continue

        if sale_price < 0 or purchase_price < 0 or quantity < 0:
            errors.append(ImportRowError(idx, "отрицательные цены/количество", snapshot))
            continue

        rows.append(
            ImportRow(
                row_index=idx,
                name=name[:255],
                barcode=barcode,
                category=category,
                sale_price=sale_price,
                purchase_price=purchase_price,
                quantity=quantity,
            )
        )

    return rows, errors
